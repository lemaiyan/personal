import json
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Load the processed data
with open(
    "/Users/lemaiyan/dev/all/personal/mum-in-law-house/expense_data.json", "r"
) as f:
    data = json.load(f)

# Create Excel workbook
wb = Workbook()

# Remove default sheet and create custom sheets
wb.remove(wb.active)

# Create sheets
summary_sheet = wb.create_sheet("Home Summary")
daily_sheet = wb.create_sheet("Daily Expenses")
category_sheet = wb.create_sheet("Category Analysis")
mpesa_sheet = wb.create_sheet("M-Pesa Fees")
outstanding_sheet = wb.create_sheet("Outstanding Balances")
unpaid_sheet = wb.create_sheet("Unpaid Labor")
pending_purchases_sheet = wb.create_sheet("Pending Purchases")

# Define styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
title_font = Font(bold=True, size=14, color="2C5F2D")
currency_font = Font(size=11)
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# === HOME SUMMARY SHEET ===
ws = summary_sheet

# Project header
ws["A1"] = "MOTHER-IN-LAW HOUSE - EXPENSE TRACKER"
ws["A1"].font = Font(bold=True, size=16, color="2C5F2D")
ws.merge_cells("A1:E1")

# Key metrics
ws["A3"] = "Project Overview"
ws["A3"].font = title_font

project_info = data["project_info"]
ws["A4"] = "Total Budget:"
ws["B4"] = f"KES {project_info['total_budget']:,}"
ws["A5"] = "Total Spent:"
ws["B5"] = f"KES {project_info['total_cost']:,}"
ws["A6"] = "Balance Remaining:"
ws["B6"] = f"KES {project_info['balance_remaining']:,}"
ws["A7"] = "Budget Used:"
ws["B7"] = f"{project_info['percentage_used']}%"
ws["A8"] = "M-Pesa Fees:"
ws["B8"] = f"KES {project_info['total_mpesa_fees']:,}"

# Outstanding amounts section
ws["A10"] = "Outstanding Amounts"
ws["A10"].font = title_font

if "total_outstanding" in project_info:
    ws["A11"] = "Outstanding Balances:"
    ws["B11"] = f"KES {project_info['total_outstanding']:,}"
    ws["A12"] = "Unpaid Labor:"
    ws["B12"] = f"KES {project_info['total_unpaid_labor']:,}"
    ws["A13"] = "Total Pending:"
    ws["B13"] = f"KES {project_info['total_pending']:,}"
    ws["A14"] = "Total Committed:"
    ws["B14"] = (
        f"KES {project_info.get('total_committed', project_info['total_cost'] + project_info['total_pending']):,}"
    )
    ws["A15"] = "Budget Used (Inclusive):"
    ws["B15"] = f"{project_info.get('percentage_used_inclusive', 0):.2f}%"
    ws["A16"] = "Effective Balance:"
    ws["B16"] = (
        f"KES {project_info.get('effective_balance', project_info['balance_remaining'] - project_info['total_pending']):,}"
    )

    category_start_row = 18

    # Add Project Completion Estimate if pending purchases exist
    if "total_pending_purchases" in project_info:
        ws["A18"] = "Project Completion Estimate"
        ws["A18"].font = title_font
        ws["A19"] = "Pending Purchases:"
        ws["B19"] = f"KES {project_info['total_pending_purchases']:,}"
        ws["A20"] = "Total Project Cost:"
        ws["B20"] = f"KES {project_info['total_project_cost']:,}"
        ws["B20"].font = Font(bold=True, size=12, color="E74C3C")
        ws["A21"] = "Additional Funds Needed:"
        ws["B21"] = f"KES {project_info['additional_funds_needed']:,}"
        ws["B21"].font = Font(bold=True, size=12, color="E74C3C")

        category_start_row = 23
else:
    category_start_row = 11

# Category summary
ws[f"A{category_start_row}"] = "Category Breakdown"
ws[f"A{category_start_row}"].font = title_font

# Headers
headers = ["Category", "Amount Spent", "M-Pesa Fees", "Total Cost", "Budget %"]
header_row = category_start_row + 1
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

# Category data
for row, category in enumerate(data["category_summary"], header_row + 1):
    ws.cell(row=row, column=1, value=category["category"]).border = border
    ws.cell(row=row, column=2, value=category["amount"]).border = border
    ws.cell(row=row, column=3, value=category["mpesa_fee"]).border = border
    ws.cell(row=row, column=4, value=category["total_cost"]).border = border
    percentage = (category["total_cost"] / project_info["total_budget"]) * 100
    ws.cell(row=row, column=5, value=f"{percentage:.2f}%").border = border

# Adjust column widths
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 12

# === DAILY EXPENSES SHEET ===
ws = daily_sheet

# Headers
expense_headers = [
    "Date",
    "Category",
    "Subcategory",
    "Description",
    "Amount",
    "M-Pesa Fee",
    "Total Cost",
    "Vendor",
    "Status",
]
for col, header in enumerate(expense_headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

# Expense data
for row, expense in enumerate(data["daily_expenses"], 2):
    ws.cell(row=row, column=1, value=expense["date"]).border = border
    ws.cell(row=row, column=2, value=expense["category"]).border = border
    ws.cell(row=row, column=3, value=expense["subcategory"]).border = border
    ws.cell(row=row, column=4, value=expense["description"]).border = border
    ws.cell(row=row, column=5, value=expense["amount"]).border = border
    ws.cell(row=row, column=6, value=expense["mpesa_fee"]).border = border
    ws.cell(row=row, column=7, value=expense["total_cost"]).border = border
    ws.cell(row=row, column=8, value=expense["vendor"]).border = border

    # Status column and highlighting for unpaid items
    status = expense.get("status", "paid")
    status_cell = ws.cell(row=row, column=9, value=status.upper())
    status_cell.border = border

    if status == "unpaid":
        # Highlight unpaid rows in light red
        unpaid_fill = PatternFill(
            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
        )
        unpaid_font = Font(color="CC0000", bold=True)
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = unpaid_fill
            if col == 9:  # Status column
                cell.font = unpaid_font

# Adjust column widths
for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
    ws.column_dimensions[col].width = 18

# === M-PESA FEES SHEET ===
ws = mpesa_sheet

ws["A1"] = "M-Pesa Fee Analysis"
ws["A1"].font = title_font

# Fee structure table
ws["A3"] = "Fee Structure"
ws["A3"].font = Font(bold=True)

fee_headers = ["Amount Range", "Fee (KES)", "Transactions", "Total Fees"]
for col, header in enumerate(fee_headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

# Calculate fee analysis from expenses
fee_analysis = {}
for expense in data["daily_expenses"]:
    amount = expense["amount"]
    fee = expense["mpesa_fee"]

    # Determine fee bracket
    if amount <= 49:
        bracket = "0-49"
    elif amount <= 100:
        bracket = "50-100"
    elif amount <= 500:
        bracket = "101-500"
    elif amount <= 1000:
        bracket = "501-1000"
    elif amount <= 1500:
        bracket = "1001-1500"
    elif amount <= 3000:
        bracket = "1501-3000"
    elif amount <= 5000:
        bracket = "3001-5000"
    elif amount <= 7500:
        bracket = "5001-7500"
    elif amount <= 10000:
        bracket = "7501-10000"
    elif amount <= 15000:
        bracket = "10001-15000"
    elif amount <= 20000:
        bracket = "15001-20000"
    elif amount <= 35000:
        bracket = "20001-35000"
    else:
        bracket = "35001+"

    if bracket not in fee_analysis:
        fee_analysis[bracket] = {"count": 0, "total_fees": 0, "fee_rate": fee}
    fee_analysis[bracket]["count"] += 1
    fee_analysis[bracket]["total_fees"] += fee

# Add fee analysis to sheet
row = 5
for bracket, analysis in fee_analysis.items():
    ws.cell(row=row, column=1, value=bracket)
    ws.cell(row=row, column=2, value=analysis["fee_rate"])
    ws.cell(row=row, column=3, value=analysis["count"])
    ws.cell(row=row, column=4, value=analysis["total_fees"])
    row += 1

# Summary
ws[f"A{row+2}"] = "Total M-Pesa Fees:"
ws[f"B{row+2}"] = f"KES {project_info['total_mpesa_fees']:,}"
ws[f"B{row+2}"].font = Font(bold=True)

# === OUTSTANDING BALANCES SHEET ===
ws = outstanding_sheet

# Title
ws["A1"] = "OUTSTANDING BALANCES"
ws["A1"].font = Font(bold=True, size=16, color="2C5F2D")
ws.merge_cells("A1:D1")

# Headers
headers = ["Vendor", "Description", "Amount", "Due Date"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

# Data
if "outstanding_balances" in data:
    for row, balance in enumerate(data["outstanding_balances"], 4):
        ws.cell(row=row, column=1, value=balance["vendor"]).border = border
        ws.cell(row=row, column=2, value=balance["description"]).border = border
        ws.cell(row=row, column=3, value=balance["amount"]).border = border
        ws.cell(row=row, column=4, value=balance["due_date"]).border = border

# Total
if "outstanding_balances" in data:
    total_row = len(data["outstanding_balances"]) + 5
    ws[f"A{total_row}"] = "Total Outstanding:"
    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"C{total_row}"] = f"KES {project_info.get('total_outstanding', 0):,}"
    ws[f"C{total_row}"].font = Font(bold=True)

# Adjust column widths
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 20

# === UNPAID LABOR SHEET ===
ws = unpaid_sheet

# Title
ws["A1"] = "UNPAID LABOR EXPENSES"
ws["A1"].font = Font(bold=True, size=16, color="2C5F2D")
ws.merge_cells("A1:D1")

# Headers
headers = ["Date", "Description", "Amount", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

# Data
if "unpaid_expenses" in data:
    for row, expense in enumerate(data["unpaid_expenses"], 4):
        ws.cell(row=row, column=1, value=expense["date"]).border = border
        ws.cell(row=row, column=2, value=expense["description"]).border = border
        ws.cell(row=row, column=3, value=expense["amount"]).border = border
        ws.cell(row=row, column=4, value="PENDING").border = border
        # Highlight pending status in red
        ws.cell(row=row, column=4).font = Font(color="E74C3C", bold=True)

# Total
if "unpaid_expenses" in data:
    total_row = len(data["unpaid_expenses"]) + 5
    ws[f"A{total_row}"] = "Total Unpaid Labor:"
    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"C{total_row}"] = f"KES {project_info.get('total_unpaid_labor', 0):,}"
    ws[f"C{total_row}"].font = Font(bold=True)

# Adjust column widths
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15

# === PENDING PURCHASES SHEET ===
ws = pending_purchases_sheet

# Title
ws["A1"] = "PENDING PURCHASES (Not Yet Procured)"
ws["A1"].font = Font(bold=True, size=16, color="2C5F2D")
ws.merge_cells("A1:C1")

# Headers
headers = ["Category", "Description", "Amount"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

# Data
if "pending_purchases" in data:
    for row, purchase in enumerate(data["pending_purchases"], 4):
        ws.cell(row=row, column=1, value=purchase["category"]).border = border
        ws.cell(row=row, column=2, value=purchase["description"]).border = border
        ws.cell(row=row, column=3, value=purchase["amount"]).border = border

        # Highlight contingency row in different color
        if "Contingency" in purchase.get("category", ""):
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
            )
            ws.cell(row=row, column=2).fill = PatternFill(
                start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
            )
            ws.cell(row=row, column=3).fill = PatternFill(
                start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
            )

    # Total
    total_row = len(data["pending_purchases"]) + 5
    ws[f"A{total_row}"] = "Total Pending Purchases:"
    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"C{total_row}"] = f"KES {project_info.get('total_pending_purchases', 0):,}"
    ws[f"C{total_row}"].font = Font(bold=True, size=12, color="E74C3C")

    # Add breakdown summary
    ws[f"A{total_row+2}"] = "SUMMARY:"
    ws[f"A{total_row+2}"].font = Font(bold=True, size=12, color="2C5F2D")
    ws[f"A{total_row+3}"] = "1. Already Spent (Paid):"
    ws[f"B{total_row+3}"] = f"KES {project_info['total_cost']:,}"
    ws[f"A{total_row+4}"] = "2. Outstanding Balances:"
    ws[f"B{total_row+4}"] = f"KES {project_info['total_outstanding']:,}"
    ws[f"A{total_row+5}"] = "3. Unpaid Labor:"
    ws[f"B{total_row+5}"] = f"KES {project_info['total_unpaid_labor']:,}"
    ws[f"A{total_row+6}"] = "4. Pending Purchases:"
    ws[f"B{total_row+6}"] = f"KES {project_info['total_pending_purchases']:,}"
    ws[f"A{total_row+8}"] = "TOTAL PROJECT COST:"
    ws[f"A{total_row+8}"].font = Font(bold=True, size=14, color="E74C3C")
    ws[f"B{total_row+8}"] = f"KES {project_info['total_project_cost']:,}"
    ws[f"B{total_row+8}"].font = Font(bold=True, size=14, color="E74C3C")
    ws[f"A{total_row+9}"] = "Current Budget:"
    ws[f"B{total_row+9}"] = f"KES {project_info['total_budget']:,}"
    ws[f"A{total_row+10}"] = "Additional Funds Needed:"
    ws[f"A{total_row+10}"].font = Font(bold=True, size=12, color="C0392B")
    ws[f"B{total_row+10}"] = f"KES {project_info['additional_funds_needed']:,}"
    ws[f"B{total_row+10}"].font = Font(bold=True, size=12, color="C0392B")

# Adjust column widths
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 15

# Save the Excel file
wb.save(
    "/Users/lemaiyan/dev/all/personal/mum-in-law-house/Mother-In-Law-House-Expenses.xlsx"
)
print("Excel file created: Mother-In-Law-House-Expenses.xlsx")
print(
    "Sheets created: Home Summary, Daily Expenses, Category Analysis, M-Pesa Fees, Outstanding Balances, Unpaid Labor, Pending Purchases"
)
